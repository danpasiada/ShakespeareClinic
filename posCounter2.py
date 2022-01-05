#filename: posCounter2.py
#Authors: Dan Pasiada with inspiration from Sam Ness, Shakespeare Clinic
#purpose: counts amounts of differents parts of speech
#tested to work in ipython -- some of packages were needed to be installed to run
import nltk # nltk has its own package installer
from nltk.tokenize import word_tokenize,sent_tokenize
from nltk import pos_tag, pos_tag_sents
from nltk.corpus import brown


from openpyxl import Workbook # excel output
from openpyxl.worksheet.table import Table, TableStyleInfo # multiple data output
from openpyxl.utils import get_column_letter # changing column width

#TODO: GUI
#TODO: look at word adjacency and custom word list #Gabriel Eagern UPenn;
####INPUT FILE NAMES BELOW######
# each file name should be in quotes and separated by a comma
# files = ['TGVVS.TXT','JEWMVS.TXT'] #edit input files here
################################

##for testing on Dan's machine##
#files = ['Shakespeare1.txt','Shakespeare2.txt']
# path = nltk.data.find('corpora/gutenberg/melville-moby_dick.txt')
# files = [path]
files = ['test.txt']

wb = Workbook()#opens new Excel workbook
ws = wb.active #makes the active worksheet our ws that we will write in
ws.title = "posCount2Results" #titles worksheet


# for deciphering tokens run:
# nltk.download('tagsets') #run this only once
# nltk.help.upenn_tagset() #this command shows all tokens, their full names and examples
# https://gist.github.com/nlothian/9240750 TODO: implement these tags

posTrack = [
        ['DT+MD'],
        ['$'],
        ['``'],
        ["''"],
        ['('],
        [')'],
        [','],
        ['--'],
        ['.'],
        [':'],
        ['CC'],
        ['CD'],
        ['DT'],
        ['EX'],
        ['FW'],
        ['IN'],
        ['JJ'],
        ['JJR'],
        ['JJS'],
        ['LS'],
        ['MD'],
        ['NN'],
        ['NNP'],
        ['NNPS'],
        ['NNS'],
        ['PDT'],
        ['POS'],
        ['PRP'],
        ['PRP$'],
        ['RB'],
        ['RBR'],
        ['RBS'],
        ['RP'],
        ['SYM'],
        ['TO'],
        ['UH'],
        ['VB'],
        ['VBD'],
        ['VBG'],
        ['VBN'],
        ['VBP'],
        ['VBZ'],
        ['WDT'],
        ['WP'],
        ['WP$'],
        ['WRB']
    ]
def makeTokens():
    '''tokenise each part of speech in each file
        return: posTrack, a nested list of each [P.O.S. tag] and its [count in file 0][count in file 1]...[count in file n]
    '''
    for fileCount,i  in enumerate(files):
        f = open(i) #current file
        raw = f.read()
        f.close()
        # posList = pos_tag(word_tokenize(raw),tagset='brown') #gives each word in the file a token, creating a very large 2D list
        # posList = pos_tag_sents(sent_tokenize(raw)) #sentence tokenisation -- currently not working
        # posList = brown.tags(brown.tagged_sents(raw)) #brown tokens -- currently not working
        posList = brown.tagged_words(brown.tagged_words(raw))
        for pos in posTrack:
            pos.append(0) #creates a new column for excel output within posTrack
            #i.e. ['CC'] -> ['CC',0]
            for token in posList:
                if token[1] == pos[0]:
                    pos[fileCount + 1] += 1

# brown_tagged_sents = brown.tagged_sents(categories='news')
# brown_sents = brown.sents(categories='news')
def changeTitles():
    '''change posTrack so that we have a human-readable name in each [P.O.S. tag] '''
    leftTitles = [
        'currently testing this',
        'dollar sign',
        'opening quotation mark',
        'closing quotation mark',
        'opening parenthesis',
        'closing parenthesis',
        'comma',
        'dash',
        'sentence terminator',
        'colon or ellipsis',
        'conjunction, coordinating',
        'numeral, cardinal',
        'determiner',
        'existential there',
        'foreign word',
        'preposition or conjunction, subordinating',
        'adjective or numeral, ordinal',
        'adjective, comparative',
        'adjective, superlative',
        'list item marker',
        'modal auxiliary',
        'noun, common, singular or mass',
        'noun, proper, singular',
        'noun, proper, plural',
        'noun, common, plural',
        'pre-determiner',
        'genitive marker',
        'pronoun, personal',
        'pronoun, possessive',
        'adverb',
        'adverb, comparative',
        'adverb, superlative',
        'particle',
        'symbol',
        '"to" as preposition or infinitive marker',
        'interjection',
        'verb, base form',
        'verb, past tense',
        'verb, present participle or gerund',
        'verb, past participle',
        'verb, present tense, not 3rd person singular',
        'verb, present tense, 3rd person singular',
        'WH-determiner',
        'WH-pronoun',
        'WH-pronoun, possessive',
        'Wh-adverb'
    ]

    for i,row in enumerate(leftTitles):
        '''adds the parts of speech on the left of the data, so we can print these instead of 'CC' and etc.'''
        posTrack[i][0] = row

def writeToFile():
    files.insert(0,'Parts of Speech')#adds empty cell to the left of file names, so that there is no filename above parts of speech

    #ws.append(files)#prints file names in the top row
    #instead, change the posTrack table to have titles on top
    posTrack.insert(0,files)
    for row in posTrack:
        ws.append(row)#writes data to worksheet


def addStyle():
    '''format data into coloured table;
        format column widths to the width of the widest cell in each column
    '''

    def colnum_string(n):
        '''convert the worksheet width to a excel-style alphabetised index'''
        string = ""
        while n > 0:
            n, remainder = divmod(n - 1, 26)
            string = chr(65 + remainder) + string
        return string

    #range from smallest to largest index in worksheet
    refValue = 'A1:' + colnum_string(ws.max_column) + '2'
    table = Table(displayName='Table1', ref=refValue)
    #make data into a table for better readability

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name='TableStyleMedium9', showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)


    #make columns as wide as their widest cell
    column_widths = []
    for row in posTrack: #get the largest width
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(str(cell)) > column_widths[i]:
                    column_widths[i] = len(str(cell))
            else:
                column_widths += [len(str(cell))]

    for i, column_width in enumerate(column_widths):
        #set the width
        ws.column_dimensions[get_column_letter(i+1)].width = column_width + 1

#'main'
makeTokens()
changeTitles()
writeToFile()
addStyle()


wb.save('posCounter2Output.xlsx')#saves workbook -- will overwrite this file
wb.close
